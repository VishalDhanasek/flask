import pandas as pd
import re
from datetime import datetime
import calendar
import os
import openpyxl
from secrets import compare_digest
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
# import flask_app as fl
import requests
import time
import threading


# Define the questions and their respective validation patterns/messages
questions_admin = [
    ("Month", "Enter Month", "^\\d{2}$", "Invalid input. Please use DD format."),
    ("Year", "Enter Year: ", "^\\d{4}$", "Invalid input. Please use YYYY format"),
    ("Project Code", "Enter Project Code: ", "^.+$", "Project code cannot be empty."),
]

shifts = ["Morning", "Afternoon", "Night"]
def employee_login(employee_id):
    try:
        # Load employee credentials from Excel file
        credentials_df = pd.read_excel("C:\\Users\\veeru\\Downloads\\roster\\credentials.xlsx")

        # Check if the employee ID exists in the DataFrame
        if str(employee_id) in credentials_df['Id'].astype(str).values:
            # Retrieve the corresponding password as string
            matching_rows = credentials_df[credentials_df['Id'].astype(str) == str(employee_id)]

            if not matching_rows.empty:
                password = str(matching_rows.iloc[0]['Password'])

                # Prompt user for password
                entered_password = input("Enter your password: ")

                # Compare entered password with the stored password
                if entered_password == password:
                    return True
                else:
                    print("Invalid password.")
                    return False
            else:
                print("No matching rows found.")
                return False
        else:
            print("Employee ID not found.")
            return False
    except FileNotFoundError:
        print("Credentials file not found.")
        return False
    except Exception as e:
        print("Error:", e)
        return False


def change_password(employee_id):
    try:
        file_path = "C:\\roster\\employee_credentials.xlsx"
        # Load employee credentials from Excel file
        credentials_df = pd.read_excel(file_path)

        # Check if the employee ID exists in the DataFrame
        if str(employee_id) in credentials_df['Employee ID'].astype(str).values:
            # Retrieve the corresponding row for the employee ID
            matching_row = credentials_df[credentials_df['Employee ID'].astype(str) == str(employee_id)]

            if not matching_row.empty:
                stored_password = str(matching_row.iloc[0]['Password'])

                # Prompt user for current password
                current_password = input("Enter your current password: ")

                # Compare entered current password with the stored password
                if current_password == stored_password:
                    # Prompt user for new password
                    new_password = input("Enter your new password: ")

                    # Update password in DataFrame
                    credentials_df.loc[credentials_df['Employee ID'].astype(str) == str(employee_id), 'Password'] = str(new_password)

                     # Debug print to check if DataFrame is updated correctly

                    # Save DataFrame to Excel file
                    credentials_df.to_excel(file_path, index=False)

                    print("Password changed successfully.")

                    return True
                else:
                    print("Incorrect current password.")
                    return False
            else:
                print("No matching rows found.")
                return False
        else:
            print("Employee ID not found.")
            return False

    except Exception as e:
        print("Error:", e)
        return False


def send_email_and_wait(employee_id_1,employee_id_2,date_employee_1,matching_employees):
    time.sleep(30)
    approval_status = None
    response = requests.get("http://localhost:5000/approval")
    if response.status_code == 200:
        approval_status = response.json().get("approval")

    else:
        print("Error fetching approval status:", response.status_code)

    # Read Excel file into a DataFrame
    df = pd.read_excel("C:\\roster\\modified_roster.xlsx")

    emp_df = pd.read_excel("C:\\roster\\employee_chatbot.xlsx")

    swap_requests_df =pd.read_excel("C:\\roster\\swap_request_log.xlsx")

    admin_df = pd.read_excel("C:\\roster\\admin.xlsx")
    month = admin_df.loc[0, "Month"]

    employee_name_1 = emp_df.loc[emp_df['employee_id'] == int(employee_id_1), "employee_name"].iloc[0]
    employee_name_2 = emp_df.loc[emp_df['employee_id'] == int(employee_id_2), "employee_name"].iloc[0]

    # Convert month number to three-letter abbreviation
    month_abbr = calendar.month_abbr[int(month)]

    # Get column indices corresponding to the provided dates
    col_date_employee_1 = f"{month_abbr} {date_employee_1}"
    col_date_employee_2 = f"{month_abbr} {date_employee_1}"

    shift_employee_1 = df.loc[df['Employee ID'] == int(employee_id_1), col_date_employee_1].iloc[0]
    shift_employee_2 = df.loc[df['Employee ID'] == int(employee_id_2), col_date_employee_1].iloc[0]
    # Record the time when the email was sent
    email_sent_time = time.time()

    # For demonstration purposes, assume a timeout period of 1 minute
    timeout_duration = 1  # 1 minute in seconds


    while time.time() - email_sent_time < timeout_duration:

        if approval_status is not None:
            if approval_status == "Accepted":
                swap_dates(employee_id_1, employee_id_2, date_employee_1)
                swap_requests_df = swap_requests_df._append({
                    "Requester_ID": employee_id_1,
                    "Requester_Name": employee_name_1,
                    "Recipient_ID": employee_id_2,
                    "Recipient_Name": employee_name_2,
                    "Date": col_date_employee_1,
                    "Shift_Of_Requester" : shift_employee_1,
                    "Requested_Shift" : shift_employee_2,
                    "Status": "Accepted"
                }, ignore_index=True)
            else:
                swap_requests_df = swap_requests_df._append({
                    "Requester_ID": employee_id_1,
                    "Requester_Name": employee_name_1,
                    "Recipient_ID": employee_id_2,
                    "Recipient_Name": employee_name_2,
                    "Date": col_date_employee_1,
                    "Shift_Of_Requester": shift_employee_1,
                    "Requested_Shift": shift_employee_2,
                    "Status": "Declined"
                }, ignore_index=True)


                # Get email addresses corresponding to employee IDs
                requested_Employee_email = emp_df.loc[emp_df['employee_id'] == int(employee_id_1), "Email ID"].iloc[0]
                email = "rotavrts@gmail.com"
                password = "rhdd gtal zuso gwnc"
                rec_email = requested_Employee_email
                subject = "Shift Swap Notification"
                # Get other employees with the same shift
                other_employees = matching_employees[(matching_employees['Employee ID'] != int(employee_id_1)) &
                                                     (matching_employees['Employee ID'] != int(employee_id_2))]
                # Prepare the message
                msg = f'Dear {employee_name_1}, \n\nThis is to inform you that your shift request has been declined on {month_abbr} - {date_employee_1} by {employee_name_2} (Employee ID: {employee_id_2})\n\nOther employees with the same shift:\n'
                for index, row in other_employees.iterrows():
                    employee_id = int(row['Employee ID'])  # Convert to integer
                    msg += f"Employee ID: {employee_id} | Employee Name: {row['Employee Name']}\n"

                msg += '\nYou can contact the above mentioned employees for shift swap.\n\nRegards,\nVirtusa'

                # Create a multipart message
                message = MIMEMultipart()
                message["From"] = email
                message["To"] = rec_email
                message["Subject"] = subject

                # Add body to email
                message.attach(MIMEText(msg, "plain"))

                # Connect to the server
                server = smtplib.SMTP("smtp.gmail.com", 587)
                server.starttls()

                # Login to the server
                server.login(email, password)

                # Send email
                server.sendmail(email, rec_email, message.as_string())
                # print("Sent email to " + rec_email)

                # print("Declined")
                # Save the updated DataFrame to an Excel file
            swap_requests_df.to_excel("C:\\roster\\swap_request_log.xlsx", index=False)

            break



def send_email_with_buttons(sender_email, receiver_email, sender_password, accept_link, decline_link,employee_id_1,employee_id_2,date_employee_1,month):
    # Read Excel file into a DataFrame
    df = pd.read_excel("C:\\roster\\modified_roster.xlsx")

    emp_df = pd.read_excel("C:\\roster\\employee_chatbot.xlsx")

    # Convert input employee IDs to integers
    employee_id_1 = int(employee_id_1)
    employee_id_2 = int(employee_id_2)

    # Locate rows corresponding to the provided employee IDs
    row_employee_1 = df[df['Employee ID'] == employee_id_1].index
    row_employee_2 = df[df['Employee ID'] == employee_id_2].index

    # Ensure both employee IDs exist in the DataFrame
    if len(row_employee_1) == 0 or len(row_employee_2) == 0:
        print("One or both of the provided employee IDs do not exist.")
        return

    # Convert month number to three-letter abbreviation
    month_abbr = calendar.month_abbr[int(month)]
    # Get column indices corresponding to the provided dates
    col_date_employee_1 = f"{month_abbr} {date_employee_1}"
    col_date_employee_2 = f"{month_abbr} {date_employee_1}"

    shift_employee_2 = df.loc[df['Employee ID'] == int(employee_id_2), col_date_employee_1].iloc[0]
    if shift_employee_2 == 'M':
        shift_employee_2 = "Morning"
    elif shift_employee_2 == 'A':
        shift_employee_2 = "Afternoon"
    elif shift_employee_2 == 'N':
        shift_employee_2 = "Night"
    else:
        shift_employee_2 = "Off"

    # Check if the provided dates exist in the DataFrame
    if col_date_employee_1 not in df.columns or col_date_employee_2 not in df.columns:
        print("One or both of the provided dates do not exist.")
        return

    employee_name_1 = emp_df.loc[emp_df['employee_id'] == int(employee_id_1), "employee_name"].iloc[0]
    employee_name_2 = emp_df.loc[emp_df['employee_id'] == int(employee_id_2), "employee_name"].iloc[0]

    # Email content
    message = MIMEMultipart("alternative")
    message["Subject"] = "Shift Swap Notification"
    message["From"] = sender_email
    message["To"] = receiver_email

    # HTML content with accept and decline buttons
    html_content = f"""
    <html>
      <body>
        <p>Dear {employee_name_2}</p>
        <p>You have a shift swap request on {col_date_employee_1}({shift_employee_2} shift) from {employee_name_1}. \nPlease review the request and respond accordingly.</p>
        <a href="{accept_link}"><button style="background-color: #4CAF50; color: white; padding: 15px 32px; text-align: center; display: inline-block; font-size: 16px;">Accept</button></a>
        <a href="{decline_link}"><button style="background-color: #f44336; color: white; padding: 15px 32px; text-align: center; display: inline-block; font-size: 16px;">Decline</button></a>
        <p>Regards,</p>
        <p>Virtusa</p>
      </body>
    </html>
    """

    # Attach HTML content to the email
    message.attach(MIMEText(html_content, "html"))

    # Connect to SMTP server and send email
    with smtplib.SMTP("smtp.gmail.com", 587) as server:
        server.starttls()
        server.login(sender_email, sender_password)
        server.sendmail(sender_email, receiver_email, message.as_string())



admins = {}

def input_login():
    username = input("Enter username: ")
    password = input("Enter password: ")
    admin_login(username, password)

def admin_login(username, password):
    if username in admins and compare_digest(admins[username], password):
        print("Login successful. Welcome, {}!".format(username))
        while True:
            choice = input("Enter your choice (1-Change Password, 2-Proceed, 3-Exit): ")

            if choice == "1":
                reset_choice = input("Change Password? (y/n): ").lower()
                if reset_choice == 'y':
                    reset_password(username)
                elif reset_choice == 'n':
                    print("Exiting...")
                    break
                else:
                    print("Invalid choice. Please enter 'y' or 'n'.")

            elif choice == "2":
                responses_admin = ask_questions_admin()
                save_to_excel_admin(responses_admin)
                print("Thank you! Your responses have been saved.")
                break

            elif choice == "3":
                print("Exiting...")
                return  # Exit the function and end the program

            else:
                print("Invalid choice.")

    else:
        print("Invalid username or password.")
        input_login()


def initial_password():
    # Create passwords file if it doesn't exist
    create_passwords_file()
    # Load passwords from Excel
    load_passwords_from_excel()

    username = input("Enter username: ")
    password = input("Enter password: ")
    if username in admins and compare_digest(admins[username], password):
        admin_login(username, password)
    else:
        print("Login Failed")

def reset_password(username):
    new_password = input("Enter new password: ")
    admins[username] = new_password
    update_password_in_excel(username, new_password)
    print("Password reset successfully. You can now login with your new password.")
    load_passwords_from_excel()  # Reload passwords after updating


def create_passwords_file():
    file_path = "C:\\roster\\admin_credentials.xlsx"
    if not os.path.isfile(file_path):
        try:
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.append(["admin", "admin"])  # Header row
            wb.save(file_path)
        except Exception as e:
            print("Error creating passwords file:", e)

def load_passwords_from_excel():
    try:
        wb = openpyxl.load_workbook("C:\\roster\\admin_credentials.xlsx")
        sheet = wb.active
        for row in sheet.iter_rows(values_only=True):
            username, password = row
            admins[username] = password
    except FileNotFoundError:
        print("Password file not found.")

def update_password_in_excel(username, new_password):
    try:
        wb = openpyxl.load_workbook("C:\\roster\\admin_credentials.xlsx")
        sheet = wb.active
        for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if row[0] == username:
                sheet.cell(row=row_index, column=2).value = new_password  # Update the password in the second column
                break
        wb.save("C:\\roster\\admin_credentials.xlsx")
    except FileNotFoundError:
        print("Password file not found.")


# Define the questions and their respective validation patterns/messages for Employees
def questions():
    questions_employees = [
        ("employee_id", "Enter employee ID (numbers only): ", "^\\d+$", "Invalid input. Please enter numbers only."),
        ("Planned_Leave_1", "Enter first planned leave (DD format) (press Enter to skip): ",
         r"^(0[1-9]|[12][0-9]|3[01])$",
         "Invalid date format. Please use DD."),
        ("Planned_Leave_2", "Enter second planned leave (DD format) (press Enter to skip): ",
         r"^(0[1-9]|[12][0-9]|3[01])$|^$",
         "Invalid date format. Please use DD or leave it empty.")
    ]
    return questions_employees


def is_valid_employee(emp_id):
    emp_id = int(emp_id)
    existing_data = pd.read_excel("C:\\roster\\employee_chatbot.xlsx",
                                  engine='openpyxl')
    employee_ids = existing_data[['employee_id']].stack().dropna().astype(int)
    for employee in employee_ids:
        # print(type(emp_id))
        # print(type(employee))
        if emp_id == employee:
            return True
    return False

def update_preferred_dates(questions_to_employees, employee_id):
    available_dates()
    print('\n')
    responses = {}
    # print(employee_id)
    if employee_id is not None:
        # Filter questions to exclude "employee_id" and include only "Planned_Leave_1" and "Planned_Leave_2"
        preferred_date_questions = [question for question in questions_to_employees if
                                    question[0] in ["Planned_Leave_1", "Planned_Leave_2"]]

        for question in preferred_date_questions:
            prompt = question[1]
            pattern = question[2]
            error_message = question[3]
            while True:
                response = input(prompt)
                if pattern and response.strip():  # Check if pattern exists and response is not empty
                    if re.match(pattern, response):
                        if question[0].startswith("Planned_Leave") and not is_valid_date(response):
                            print("Invalid date. Please enter a valid day of the month.")
                        else:
                            # print('Hello')
                            responses['employee_id'] = employee_id  # Assign employee id to the employee id
                            responses[question[0]] = response  # Assign response to the question identifier
                            break
                    else:
                        print(error_message)
                else:
                    # Set response to None if input is empty
                    responses['employee_id'] = employee_id
                    responses[question[0]] = None
                    break
    return responses


# Function to ask questions and validate responses for admin
def ask_questions_admin():
    responses = {}
    for question, prompt, pattern, error_message in questions_admin:
        while True:
            response = input(prompt)
            if pattern:
                if re.match(pattern, response):
                    responses[question] = response
                    break
                else:
                    print(error_message)
            else:
                responses[question] = response
                break

    # Get the first day of the month and its corresponding day of the week
    first_day_of_month = datetime(int(responses["Year"]), int(responses["Month"]), 1)

    for shift in shifts:
        shift_demand = []
        for day in range(1, 8):
            # Get the day of the week for the current day
            current_day_of_week = calendar.day_name[(first_day_of_month.weekday() + day - 1) % 7]
            demand = input(f"What is the demand for {shift} shift on {current_day_of_week}? ")
            while not re.match("^[1-9]$", demand):
                print("Invalid demand format. Please enter a number")
                demand = input(f"What is the demand for {shift} shift on {current_day_of_week}? ")
            shift_demand.append(demand)
        responses[shift] = shift_demand

    return responses


def save_to_excel_admin(responses):
    try:
        existing_data = pd.read_excel("C:\\roster\\admin.xlsx")
        responses['Month'] = responses['Month'].zfill(2)

        # Update the first 7 rows of 'existing_data' with the data from 'responses'
        for i in range(7):
            for col in existing_data.columns:
                if col in responses:
                    value = responses[col]
                    # Check if the value is a list and handle it accordingly
                    if isinstance(value, list):
                        # Update with the first element of the list
                        value = value[0] if value else None
                    # Explicitly cast the value to the correct data type
                    if pd.api.types.is_numeric_dtype(existing_data[col]):
                        value = float(value) if '.' in str(value) else int(value)
                    existing_data.at[i, col] = value

        print("Admin data updated.")

        # Write the updated data back to the Excel file
        existing_data.to_excel("C:\\roster\\admin.xlsx", index=False)
    except FileNotFoundError:
        # If the file doesn't exist, create a new DataFrame with the new responses and write it to the Excel file
        df = pd.DataFrame([responses])
        df.to_excel("C:\\roster\\admin.xlsx", index=False)
        print("New admin data added.")


# Function to validate if the entered date is valid within a month
def is_valid_date(day):
    try:
        day = int(day)
        year = datetime.now().year
        month = datetime.now().month
        num_days = calendar.monthrange(year, month)[1]
        return 1 <= day <= num_days
    except ValueError:
        return False


# Function to ask questions and validate responses for employees
def ask_questions_employees(questions_employees):
    responses = {}
    for question, prompt, pattern, error_message in questions_employees:
        while True:
            response = input(prompt)
            if pattern:
                if response.strip():
                    if re.match(pattern, response):
                        if question.startswith("Preferred_Date") and not is_valid_date(response):
                            print("Invalid date. Please enter a valid day of the month.")
                        else:
                            responses[question] = response
                            break
                    else:
                        print(error_message)
                else:
                    # Set response to None if input is empty
                    responses[question] = None
                    break
            else:
                responses[question] = response
                break
    return responses


def available_dates():
    leave_request = {}
    existing_data = pd.read_excel("C:\\roster\\employee_chatbot.xlsx",
                                  engine='openpyxl')
    dates = existing_data[['Planned_Leave_1', 'Planned_Leave_2']].stack().dropna().astype(int)

    # Get the current year and month
    current_year = datetime.now().year
    current_month = datetime.now().month

    # Get the number of days in the current month
    num_days = calendar.monthrange(current_year, current_month)[1]

    # Generate all dates of the current month
    dates_of_current_month = [datetime(current_year, current_month, day) for day in range(1, num_days + 1)]

    print("Dates Available")
    for date in dates_of_current_month:
        day_of_month = date.day
        leave_request[day_of_month] = 0
        for iterator in dates:
            if iterator == day_of_month:
                leave_request[day_of_month] += 1

        if leave_request[day_of_month] <= 1:
            # print(f"{date},")
            print(f"{day_of_month}", end="   ")
    print('/n')
    print("Dates Taken")
    for date in dates_of_current_month:
        day_of_month = date.day
        leave_request[day_of_month] = 0
        for iterator in dates:
            if iterator == day_of_month:
                leave_request[day_of_month] += 1

        if leave_request[day_of_month] > 1:
            # print(f"{date},")
            print(f"{day_of_month}", end="   ")

def save_to_excel_employees(responses):
    leave_request = {}
    try:
        # Read existing data from Excel file if it exists
        existing_data = pd.read_excel("C:\\roster\\employee_chatbot.xlsx",
                                      engine='openpyxl')

        employee_id_store = None
        # Convert existing employee_id values to integers for comparison
        existing_data['employee_id'] = existing_data['employee_id'].astype(int)

        # Convert responses employee_id to integer for comparison
        responses['employee_id'] = int(responses['employee_id'])

        # Check if the preferred dates are already taken
        preferred_dates = [responses['Planned_Leave_1'], responses['Planned_Leave_2']]
        if all(date is None for date in preferred_dates):
            print("Both preferred dates are None")
        elif any(date is None for date in preferred_dates):
            print("At least one preferred date is None")
        else:
            # Convert string representations to actual values using eval
            # Convert string representations to actual integers
            preferred_dates = [int(date) for date in preferred_dates]

        if responses['employee_id'] in existing_data['employee_id'].values:
            # print(preferred_dates, preferred_dates[0], type(preferred_dates[0]), existing_data['Preferred_Date_1'],
            # type(existing_data['Preferred_Date_1']))
            dates = existing_data[['Planned_Leave_1', 'Planned_Leave_2']].stack().dropna().astype(int)
            # Print all dates of the current month
            for date in preferred_dates:
                # print('Hi')
                date_fixed = False
                for iterator in dates:
                    # print('Hello')
                    if iterator == date:
                        if iterator not in leave_request:
                            leave_request[iterator] = 0

                        leave_request[iterator] += 1
                        # print(iterator, leave_request[iterator])

                        if leave_request[iterator] >= 2:
                            print('\n')
                            employee_id_store = responses['employee_id']
                            date_fixed = True
                            # break
                        # if leave_request[iterator] > 2:
                        #     print(f"Preferred date {date} is already taken. Please choose another date.")
                        #     available_dates()
                if date_fixed:
                    employee_id_store = responses['employee_id']
                    break
        if employee_id_store is None:
            existing_data.loc[existing_data['employee_id'] == responses['employee_id'], 'Planned_Leave_1'] = \
                responses['Planned_Leave_1']
            existing_data.loc[existing_data['employee_id'] == responses['employee_id'], 'Planned_Leave_2'] = \
                responses['Planned_Leave_2']
            # Write the combined data to the Excel file
            existing_data.to_excel("C:\\roster\\employee_chatbot.xlsx", index=False,
                                   engine='openpyxl')
        return employee_id_store
    except FileNotFoundError:
        print("Employee Data Doesn't exist")

# Define the output path where the modified Excel file should be saved
output_path = "C:\\roster\\modified_roster.xlsx"

def swap_dates(employee_id_1, employee_id_2, date_employee_1):
    # Read Excel file into a DataFrame
    df = pd.read_excel("C:\\roster\\modified_roster.xlsx")

    emp_df=pd.read_excel("C:\\roster\\employee_chatbot.xlsx")

    admin_df=pd.read_excel("C:\\roster\\admin.xlsx")

    month = admin_df.loc[0, "Month"]
    year = admin_df.loc[0,"Year"]

    # Convert input employee IDs to integers
    employee_id_1 = int(employee_id_1)
    employee_id_2 = int(employee_id_2)

    # Locate rows corresponding to the provided employee IDs
    row_employee_1 = df[df['Employee ID'] == employee_id_1].index
    row_employee_2 = df[df['Employee ID'] == employee_id_2].index

    # Get email addresses corresponding to employee IDs
    sender_email = emp_df.loc[emp_df['employee_id'] == int(employee_id_1), "Email ID"].iloc[0]
    email_2=emp_df.loc[emp_df['employee_id'] == int(employee_id_2), "Email ID"].iloc[0]
    employee_name_1 =emp_df.loc[emp_df['employee_id'] == int(employee_id_1), "employee_name"].iloc[0]
    employee_name_2 = emp_df.loc[emp_df['employee_id'] == int(employee_id_2), "employee_name"].iloc[0]

    # Ensure both employee IDs exist in the DataFrame
    if len(row_employee_1) == 0 or len(row_employee_2) == 0:
        print("One or both of the provided employee IDs do not exist.")
        return

    # Convert month number to three-letter abbreviation
    month_abbr = calendar.month_abbr[int(month)]

    # Get column indices corresponding to the provided dates
    col_date_employee_1 = f"{month_abbr} {date_employee_1}"
    col_date_employee_2 = f"{month_abbr} {date_employee_1}"

    # Check if the provided dates exist in the DataFrame
    if col_date_employee_1 not in df.columns or col_date_employee_2 not in df.columns:
        print("One or both of the provided dates do not exist.")
        return

    # Swap values for the specified dates
    temp_value = df.at[row_employee_1[0], col_date_employee_1]
    df.at[row_employee_1[0], col_date_employee_1] = df.at[row_employee_2[0], col_date_employee_2]
    df.at[row_employee_2[0], col_date_employee_2] = temp_value

    # Write the modified DataFrame to the specified output Excel file
    with pd.ExcelWriter(output_path, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False)

    send_email_to_cab_service(month_abbr,date_employee_1,employee_name_1,employee_id_1,employee_id_2,employee_name_2,year,sender_email)
    send_email_to_manager(month_abbr,date_employee_1,employee_name_1,employee_id_1,employee_id_2,employee_name_2,year)

    email = "rotavrts@gmail.com"
    password = "rhdd gtal zuso gwnc"

    subject = "Shift Swap Notification"
    msg = f'Dear {employee_name_2}, \n\nThis is to inform you that your shift has been swapped on {month_abbr} - {date_employee_1} by {employee_name_1}(Employee ID: {employee_id_1})\n\nRegards,\nVirtusa'

    # Create a multipart message
    message = MIMEMultipart()
    message["From"] = email
    message["To"] = email_2

    message["Subject"] = subject

    # Add body to email
    message.attach(MIMEText(msg, "plain"))

    # Connect to the server
    server = smtplib.SMTP("smtp.gmail.com", 587)
    server.starttls()

    # Login to the server
    server.login(email, password)


    # Send email
    server.sendmail(email, email_2, message.as_string())

    send_mail_to_requester(month_abbr,date_employee_1,employee_name_1,employee_id_1,employee_id_2,employee_name_2,year,sender_email)
    # print("Sent email to " + rec_email)
    #
    # print("Values swapped successfully. Modified data saved to", output_path)

def send_email_to_manager(month_abbr,date_employee_1,employee_name_1,employee_id_1,employee_id_2,employee_name_2,year):
    df = pd.read_excel("C:\\roster\\modified_roster.xlsx")
    # Get column indices corresponding to the provided dates
    col_date_employee_1 = f"{month_abbr} {date_employee_1}"
    col_date_employee_2 = f"{month_abbr} {date_employee_1}"
    shift_employee_2 = df.loc[df['Employee ID'] == int(employee_id_2), col_date_employee_1].iloc[0]
    if shift_employee_2 == 'M':
        shift_employee_2 = "Morning"
    elif shift_employee_2 == 'A':
        shift_employee_2 = "Afternoon"
    elif shift_employee_2 == 'N':
        shift_employee_2 = "Night"
    else:
        shift_employee_2 = "Off"
    email = "rotavrts@gmail.com"
    password = "rhdd gtal zuso gwnc"
    manager_email = "madheshns57@gmail.com"

    subject = "Shift Swap Notification"
    msg = f'Dear Madhesh, \n\nThis is to inform you that {employee_name_1} (Employee ID: {employee_id_1}) has swapped the shift with {employee_name_2} (Employee ID: {employee_id_2}) on {col_date_employee_1} , {year}\n\nRegards,\nVirtusa'

    # Create a multipart message
    message = MIMEMultipart()
    message["From"] = email
    message["To"] = manager_email
    message["Subject"] = subject

    # Add body to email
    message.attach(MIMEText(msg, "plain"))

    # Connect to the server
    server = smtplib.SMTP("smtp.gmail.com", 587)
    server.starttls()

    # Login to the server
    server.login(email, password)

    # Send email
    server.sendmail(email, manager_email, message.as_string())
    server.quit()
def send_mail_to_requester(month_abbr,date_employee_1,employee_name_1,employee_id_1,employee_id_2,employee_name_2,year,sender_email):
    df = pd.read_excel("C:\\roster\\modified_roster.xlsx")
    # Get column indices corresponding to the provided dates
    col_date_employee_1 = f"{month_abbr} {date_employee_1}"
    col_date_employee_2 = f"{month_abbr} {date_employee_1}"
    shift_employee_1 = df.loc[df['Employee ID'] == int(employee_id_1), col_date_employee_1].iloc[0]
    if shift_employee_1 == 'M':
        shift_employee_1 = "Morning"
    elif shift_employee_1 == 'A':
        shift_employee_1 = "Afternoon"
    elif shift_employee_1 == 'N':
        shift_employee_1 = "Night"
    else:
        shift_employee_1 = "Off"
    email = "rotavrts@gmail.com"
    password = "rhdd gtal zuso gwnc"

    subject = "Shift Swap Notification"
    msg = f'Dear {employee_name_1}, \n\nThis is to inform you that the {shift_employee_1} shift you have requested on {month_abbr} {date_employee_1} {year}, has been approved by {employee_name_2}(Employee id {employee_id_2}).\n\nRegards,\nVirtusa'

    # Create a multipart message
    message = MIMEMultipart()
    message["From"] = email
    message["To"] = sender_email
    message["Subject"] = subject

    # Add body to email
    message.attach(MIMEText(msg, "plain"))

    # Connect to the server
    server = smtplib.SMTP("smtp.gmail.com", 587)
    server.starttls()

    # Login to the server
    server.login(email, password)

    # Send email
    server.sendmail(email, sender_email, message.as_string())
    server.quit()

def send_email_to_cab_service(month_abbr,date_employee_1,employee_name_1,employee_id_1,employee_id_2,employee_name_2,year,sender_email):
    df = pd.read_excel("C:\\roster\\modified_roster.xlsx")
    # Get column indices corresponding to the provided dates
    col_date_employee_1 = f"{month_abbr} {date_employee_1}"
    col_date_employee_2 = f"{month_abbr} {date_employee_1}"
    shift_employee_2 = df.loc[df['Employee ID'] == int(employee_id_2), col_date_employee_1].iloc[0]
    if shift_employee_2=='M':
        shift_employee_2="Morning"
    elif shift_employee_2=='A':
        shift_employee_2="Afternoon"
    elif shift_employee_2 =='N':
        shift_employee_2="Night"
    else:
        shift_employee_2="Off"
    email = "rotavrts@gmail.com"
    password = "rhdd gtal zuso gwnc"
    cab_email="cabvrts@gmail.com"
    subject = "Shift Swap Notification for Cab Service"
    if shift_employee_2=="Off":
        msg=f'Dear Cab Service Provider, \n\nThis is to inform you that {employee_name_2} with Employee ID {employee_id_2} will be on leave on {month_abbr} {date_employee_1}, {year}. Kindly cancel the cab facility for the above mentioned day. \n\nRegards,\nVirtusa'
    else:
        msg = f'Dear Cab Service Provider, \n\nThis is to inform you that the shift has been updated to {shift_employee_2} shift for the employee {employee_name_2} with the ID {employee_id_2} on {month_abbr} {date_employee_1}, {year}. Kindly make necessary arrangements for the swap.\n\nRegards,\nVirtusa'

    # Create a multipart message
    message = MIMEMultipart()
    message["From"] = email
    message["To"] = cab_email
    message["Subject"] = subject

    # Add body to email
    message.attach(MIMEText(msg, "plain"))

    # Connect to the server
    server = smtplib.SMTP("smtp.gmail.com", 587)
    server.starttls()

    # Login to the server
    server.login(email, password)

    # Send email
    server.sendmail(email, cab_email, message.as_string())
    server.quit()

def main():
    print("Welcome to the Chatbot!")
    global count  # Declare count as a global variable

    while True:
        choice = input("Enter your choice (1-Admin, 2-User, 3-Swap, or 4-Exit): ")

        if choice == "1":
            while True:
                choice = input("Would you like to login as admin? (y/n): ").lower()
                if choice == 'y':
                    initial_password()
                    break
                elif choice == 'n':
                    print("Exiting...")
                    break
                else:
                    print("Invalid choice. Please enter 'y' or 'n'.")

        elif choice == "2":

            employee_id = input("Enter your Employee ID: ")
            if employee_login(employee_id):
                while True:
                    user_choice = input(
                        "Enter your choice (1-Change Password, 2- Planned Leave, 3-Exit): ")
                    if user_choice == "1":
                        if change_password(employee_id):
                            print("Please log in again with your new password.")
                            break  # Break the inner loop and prompt for login again
                        else:
                            print("Password change failed.")
                    elif user_choice == "2":
                        questions_to_employees = questions()
                        responses_employees = ask_questions_employees(questions_to_employees)
                        employee_id = save_to_excel_employees(responses_employees)
                        while employee_id is not None:
                            questions_to_employees = questions()
                            responses_employees = update_preferred_dates(questions_to_employees, employee_id)
                            employee_id = save_to_excel_employees(responses_employees)
                        save_to_excel_employees(responses_employees)
                    elif user_choice == "3":
                        print("Exiting...")
                        break  # Exit the inner loop and go back to the main menu
                    else:
                        print("Invalid choice!")
        elif choice == "3":
                employee_id_1 = input("Enter Employee ID 1: ")
                employee_id_1=int(employee_id_1)
                if employee_login(employee_id_1):
                    while True:
                        user_choice = input(
                            "Enter your choice (1-Change Password, 2-Swap your shift, 3-Exit): ")
                        if user_choice == "1":
                            if change_password(employee_id_1):
                                print("Please log in again with your new password.")
                                break  # Break the inner loop and prompt for login again
                            else:
                                print("Password change failed.")
                        elif user_choice =="2":
                            date_employee_1 = input("Enter Date: ")
                            # Read Excel file into a DataFrame
                            df = pd.read_excel("C:\\Users\\veeru\\Downloads\\roster\\modified_roster.xlsx")
                            admin_df = pd.read_excel("C:\\Users\\veeru\\Downloads\\roster\\admin.xlsx")
                            month = admin_df.loc[0, "Month"]

                            # Convert month number to three-letter abbreviation
                            month_abbr = calendar.month_abbr[int(month)]

                            # Get column indices corresponding to the provided dates
                            col_date_employee_1 = f"{month_abbr} {date_employee_1}"
                            col_date_employee_2 = f"{month_abbr} {date_employee_1}"

                            # Check if the provided dates exist in the DataFrame
                            if col_date_employee_1 not in df.columns or col_date_employee_2 not in df.columns:
                                print("One or both of the provided dates do not exist.")
                                return

                            shift_to_swap=input("Which shift do you prefer?(M/A/N/G/O)").upper()

                            # Filter the DataFrame based on the provided date and shift
                            matching_employees = df[(df[col_date_employee_1] == shift_to_swap)]
                            # Print the names and IDs of matching employees
                            if not matching_employees.empty:
                                print("Employees with the specified shift on the given date:")
                                for index, row in matching_employees.iterrows():
                                    employee_id = int(row['Employee ID'])
                                    print("Employee ID:",  employee_id, " | Employee Name:", row['Employee Name'])
                            else:
                                print("No employees found with the specified shift on the given date.")

                            employee_id_2 = input("Enter Employee ID from the above employees: ")

                            sender_email = "rotavrts@gmail.com"
                            sender_password = "rhdd gtal zuso gwnc"
                            accept_link = "http://localhost:5000/accept"
                            decline_link = "http://localhost:5000/decline"



                            # Read Excel file into a DataFrame
                            df = pd.read_excel("C:\\roster\\modified_roster.xlsx")

                            emp_df = pd.read_excel("C:\\roster\\employee_chatbot.xlsx")

                            # Convert input employee IDs to integers
                            employee_id_1 = int(employee_id_1)
                            employee_id_2 = int(employee_id_2)

                            # Locate rows corresponding to the provided employee IDs
                            row_employee_1 = df[df['Employee ID'] == employee_id_1].index
                            row_employee_2 = df[df['Employee ID'] == employee_id_2].index

                            # Ensure both employee IDs exist in the DataFrame
                            if len(row_employee_1) == 0 or len(row_employee_2) == 0:
                                print("One or both of the provided employee IDs do not exist.")
                                return

                            # Get email addresses corresponding to employee IDs
                            receiver_email = emp_df.loc[emp_df['employee_id'] == int(employee_id_2), "Email ID"].iloc[0]
                            employee_name_1 = emp_df.loc[emp_df['employee_id'] == int(employee_id_1), "employee_name"].iloc[0]
                            employee_name_2 = emp_df.loc[emp_df['employee_id'] == int(employee_id_2), "employee_name"].iloc[0]

                            send_email_with_buttons(sender_email, receiver_email,sender_password, accept_link, decline_link,employee_id_1,employee_id_2,date_employee_1,month)



                            # Start a new thread to send email and wait for response
                            email_thread = threading.Thread(target=send_email_and_wait,
                                                args=(employee_id_1, employee_id_2, date_employee_1,matching_employees))
                            email_thread.start()

                            # Continue running the chatbot without waiting for the email response
                            print("Email sent. Waiting for response in the background.")

                        elif user_choice == "3":
                            print("Exiting...")
                            break  # Exit the inner loop and go back to the main menu
                        else:
                            print("Invalid choice!")

        elif choice == "4":
            print("Exiting the Chatbot.")
            break
        else:
            print("Invalid choice!")

if __name__ == "__main__":
    main()